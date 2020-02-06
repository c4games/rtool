#coding=utf-8

import xlrd
import json
import os
import sys
import yaml
import re
import click

import openpyxl

from rtool import utils
from rtool.pcutils import cache_tool
from rtool.jobs.config.CommonReadXlsJob import CommonReadXlsJob

class ReadXlsJob(CommonReadXlsJob):
    def __init__(self, options):
        CommonReadXlsJob.__init__(self, options)
        self.Save_None_Value = True

    # def get_field_list(self, worksheet):
    #     temp_field_type_list, field_name_list, filter_info = CommonReadXlsJob.get_field_list(self, worksheet)
    #     field_type_list = []
    #     for i in range(len(temp_field_type_list)):
    #         field_type_list.append(self.value_parser.format_field(temp_field_type_list[i]))
    #     return field_type_list, field_name_list, filter_info



    # @property
    # def value_parser(self):
    #     if not hasattr(self, '_value_parser'):
    #         from rtool.projects.demo.config.ValueParser import ValueParser
    #         self._value_parser = ValueParser(self._options)
    #
    #     return self._value_parser

    @property
    def filter_row_num(self):
        return 2

    @property
    def title_row_num(self):
        return 4

    @property
    def type_row_num(self):
        return 3

    def start_column_num(self, worksheet):
        return 1

    def end_column_num(self, worksheet):
        return worksheet.max_column

    def start_row_num(self, worksheet):
        return 6

    def end_row_num(self, worksheet):
        for i in range(worksheet.max_row)[::-1]:
            v = self.read_cell_value(worksheet, i, 0)
            if v == 'end' or v == 'End':
                return i + 1
        return worksheet.max_row

    def get_primary_col(self, worksheet):
        return 1
